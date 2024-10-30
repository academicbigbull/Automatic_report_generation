#操作excel表格的库openyxl  输出获取到的文档的表名
from openpyxl import load_workbook
workbook = load_workbook(filename = "E:/python/操作文档/test.xlsx")
print(workbook.sheetnames)

#通过sheet名称获取表格
sheet=workbook['Sheet1']
print(sheet)

#获取表格的尺寸  输出A1：C5表示列是A到C每一列有5个元素
print(sheet.dimensions)

#获取表格的具体数据
# workbook.active 打开激活的表格；
# sheet["A1"] 获取A1格子的数据；
# cell.value 获取格子中的值；
sheet=workbook.active
print(sheet)
cell1=sheet['A1']  #A列第一个数据
cell2=sheet['C3']  #C列第三个数据
print(cell1.value,cell2.value)
cell3=sheet.cell(row=1,column=1)
cell4=sheet.cell(row=3,column=3)
print(cell3.value,cell4.value)

#获取格子的行列以及坐标
print(cell3.row,cell3.column,cell3.coordinate)

#获取一系列的格子
cell5=sheet['A1:C3']
print(cell5)
#拿具体的值
for i in cell5:
    for j in i:
        print(j.value)

# sheet["A"] - -- 获取A列的数据
# sheet["A:C"] - -- 获取A, B, C三列的数据
# sheet[5] - -- 只获取第5行的数据


#sheet。rows（）获取所有的行
for i in sheet.rows:
    print(i)